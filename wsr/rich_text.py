"""Preserve Excel rich-text strikethrough when filling PPT tables."""

from __future__ import annotations

from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from pptx.oxml.ns import qn

# (text, strikethrough?)
TextRun = tuple[str, bool]


def runs_from_value(value) -> list[TextRun]:
    """Convert a cell value (plain, rich, or run list) into display runs."""
    if value is None:
        return [("", False)]
    # CellRichText subclasses list — check it before the generic list branch.
    if isinstance(value, CellRichText):
        runs: list[TextRun] = []
        for block in value:
            if isinstance(block, TextBlock):
                text = block.text or ""
                strike = bool(block.font and block.font.strike)
                if text:
                    runs.append((text, strike))
            else:
                text = str(block)
                if text:
                    runs.append((text, False))
        return runs or [("", False)]
    if isinstance(value, list):
        out: list[TextRun] = []
        for item in value:
            if isinstance(item, tuple) and len(item) == 2:
                out.append((str(item[0]), bool(item[1])))
            else:
                text = str(item)
                if text:
                    out.append((text, False))
        return out or [("", False)]
    try:
        import pandas as pd

        if isinstance(value, float) and pd.isna(value):
            return [("", False)]
        if value is pd.NA:
            return [("", False)]
    except Exception:
        pass
    text = str(value)
    if text.lower() in ("nan", "none", "nat"):
        return [("", False)]
    return [(text, False)]


def runs_to_plain(runs: Iterable[TextRun]) -> str:
    return "".join(text for text, _ in runs)


def full_cell_text(value) -> str:
    """Plain display text including every line / struck segment."""
    plain = runs_to_plain(runs_from_value(value)).replace("_x000B_", "\n")
    plain = plain.replace("\r\n", "\n").replace("\r", "\n").strip()
    if plain.lower() in ("nan", "none", "nat"):
        return ""
    return plain


def build_rich_run_index(path: str | Path, sheet_name: str) -> dict[tuple[int, str], list[TextRun]]:
    """
    Map (Excel 1-based row, column header) -> runs for cells that contain
    strikethrough. Callers fall back to the dataframe value otherwise.
    """
    path = Path(path)
    wb = load_workbook(path, data_only=False, rich_text=True, keep_vba=False)
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    headers: dict[int, str] = {}
    for col in range(1, (ws.max_column or 1) + 1):
        header = ws.cell(1, col).value
        if header is not None:
            headers[col] = str(header)

    index: dict[tuple[int, str], list[TextRun]] = {}
    for row in range(2, (ws.max_row or 1) + 1):
        for col, header in headers.items():
            value = ws.cell(row, col).value
            if not isinstance(value, CellRichText):
                continue
            runs = runs_from_value(value)
            if any(strike for _, strike in runs):
                index[(row, header)] = runs
    return index


def _set_run_strike(run, strike: bool) -> None:
    """Set DrawingML strikethrough via a:rPr/@strike (not a child element)."""
    r_pr = run._r.get_or_add_rPr()
    # Remove any invalid child <a:strike> from earlier attempts.
    for child in list(r_pr):
        if child.tag == qn("a:strike"):
            r_pr.remove(child)
    if strike:
        r_pr.set("strike", "sngStrike")
    elif r_pr.get("strike") is not None:
        del r_pr.attrib["strike"]


def _flatten_runs_to_lines(runs: list[TextRun]) -> list[TextRun]:
    """One entry per visual line; newline starts a new line. Drop blank lines."""
    lines: list[TextRun] = []
    current_text: list[str] = []
    current_strike = False

    def flush() -> None:
        nonlocal current_text, current_strike
        text = "".join(current_text)
        if text.strip():
            lines.append((text.strip(), current_strike))
        current_text = []

    for text, strike in runs:
        pieces = (text or "").replace("_x000B_", "\n").split("\n")
        for idx, piece in enumerate(pieces):
            if idx > 0:
                flush()
                current_strike = strike
            if not piece:
                continue
            if not current_text:
                current_strike = strike
            if current_text and strike != current_strike:
                flush()
                current_strike = strike
            current_text.append(piece)
            current_strike = strike
    flush()
    return lines or [("", False)]


def format_date_piece(text: str) -> str:
    """Format one date fragment; keep day-month-only text as-is."""
    from wsr.tracker import format_date_fragment

    return format_date_fragment(text)


def format_date_runs(runs: list[TextRun]) -> list[TextRun]:
    """Format each line of a rich date cell; keep strike flags."""
    formatted: list[TextRun] = []
    for text, strike in runs:
        pieces = [part.strip() for part in text.replace("_x000B_", "\n").split("\n")]
        for idx, part in enumerate(pieces):
            if idx > 0:
                formatted.append(("\n", False))
            if not part:
                continue
            formatted.append((format_date_piece(part), strike))
    while formatted and formatted[0][0] == "\n":
        formatted.pop(0)
    while formatted and formatted[-1][0] == "\n":
        formatted.pop()
    return formatted or [("-", False)]


def resolve_display(
    value,
    *,
    excel_row: int | None = None,
    column: str | None = None,
    rich_runs: dict[tuple[int, str], list[TextRun]] | None = None,
    as_date: bool = False,
):
    """
    Prefer indexed rich-text runs (with strikethrough) when available.
    Otherwise keep the full plain cell text, including every line.
    """
    if rich_runs and excel_row is not None and column:
        runs = rich_runs.get((excel_row, column))
        if runs:
            return format_date_runs(runs) if as_date else runs

    text = full_cell_text(value)
    if as_date:
        if not text:
            return "-"
        if "\n" in text:
            return "\n".join(
                format_date_piece(part) for part in text.split("\n") if part.strip()
            )
        return format_date_piece(text)

    import math

    try:
        if isinstance(value, float) and math.isnan(value):
            return ""
        if isinstance(value, float) and value == int(value):
            return str(int(value))
    except (TypeError, ValueError, OverflowError):
        pass
    return text


def write_table_cell(cell, value) -> list[TextRun]:
    """
    Write plain text or rich runs into a PPT table cell.

    Returns line-level runs written (for re-applying strikethrough after styling).
    """
    runs = runs_from_value(value)
    lines = _flatten_runs_to_lines(runs)

    cell.text = lines[0][0] if lines else ""
    tf = cell.text_frame
    paragraphs = list(tf.paragraphs)
    while len(paragraphs) < len(lines):
        tf.add_paragraph()
        paragraphs = list(tf.paragraphs)

    for line_idx, (text, strike) in enumerate(lines):
        paragraph = paragraphs[line_idx]
        for run in list(paragraph.runs):
            run.text = ""
        if paragraph.runs:
            run = paragraph.runs[0]
        else:
            run = paragraph.add_run()
        run.text = text
        _set_run_strike(run, strike)

    return lines


def apply_cell_strikes(cell, lines: list[TextRun]) -> None:
    """Re-apply strikethrough after style_table_cells resets run fonts."""
    run_list = []
    for paragraph in cell.text_frame.paragraphs:
        run_list.extend(paragraph.runs)
    for run, (_, strike) in zip(run_list, lines):
        _set_run_strike(run, strike)
