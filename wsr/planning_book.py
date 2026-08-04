"""
Quarterly planning metrics for the planning slide.

Bar 1 — Q3 Actual Available hours from Book2.xlsx
  (row labelled "Total work Hrs. Available for PFS team").

Bar 2 — Sum of Scrum Non STLA effort hours:
  - Prefer ``Total Revised Estimation`` (or legacy ``Revised Estimation``)
    when that cell has a value.
  - Otherwise use ``Estimated Hrs``.
  - If no revised column exists, sum Estimated Hrs only.

Bar 3 — Burndown:
  sum(Actual Efforts (Hrs)) − sum(Competency Gap Efforts).
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from wsr.constants import DEFAULT_PLANNING_BOOK

# Kept for CLI/API compatibility; no longer drives the second planning bar.
DEFAULT_PLANNED_BANDWIDTH_PCT = 90
PLANNED_BANDWIDTH_PCT = DEFAULT_PLANNED_BANDWIDTH_PCT

_AVAILABLE_LABEL = "total work hrs available for pfs team"

COL_ESTIMATED_HRS = "Estimated Hrs"
# Preferred / legacy names — actual header is resolved at runtime.
COL_REVISED_ESTIMATION = "Total Revised Estimation"
_REVISED_ESTIMATION_ALIASES = (
    "total revised estimation",
    "revised estimation",
)
COL_ACTUAL_EFFORTS = "Actual Efforts (Hrs)"
COL_COMPETENCY_GAP = "Competency Gap Efforts"


def _as_float(value) -> float | None:
    try:
        if value in (None, ""):
            return None
        if isinstance(value, float) and value != value:  # NaN
            return None
        return float(value)
    except (TypeError, ValueError):
        return None


def _normalize_header(value) -> str:
    return " ".join(str(value).strip().lower().replace(".", " ").split())


def revised_estimation_column(columns) -> str | None:
    """Resolve Total Revised Estimation / Revised Estimation despite header typos."""
    normalized = {_normalize_header(col): col for col in columns}
    for alias in _REVISED_ESTIMATION_ALIASES:
        if alias in normalized:
            return normalized[alias]
    for key, original in normalized.items():
        if "revised" in key and "estimation" in key:
            return original
    return None


def _normalize_label(value) -> str:
    return " ".join(str(value).strip().lower().replace(".", " ").split())


def _lookup_available_hours(ws) -> tuple[float | None, int | None]:
    candidates: list[dict] = []
    for row_idx in range(1, (ws.max_row or 0) + 1):
        label = ws.cell(row_idx, 2).value
        if not label or _AVAILABLE_LABEL not in _normalize_label(label):
            continue
        hours = _as_float(ws.cell(row_idx, 4).value)
        if hours is None:
            continue
        members = _as_float(ws.cell(row_idx, 9).value)
        candidates.append(
            {
                "hours": hours,
                "members": int(round(members)) if members is not None else None,
            }
        )

    if not candidates:
        return None, None

    with_members = [c for c in candidates if c["members"] is not None]
    if len(with_members) >= 2:
        max_members = max(c["members"] for c in with_members)
        subsets = [c for c in with_members if c["members"] < max_members]
        if subsets:
            best = max(subsets, key=lambda c: c["hours"])
            return best["hours"], best["members"]

    best = with_members[-1] if with_members else candidates[-1]
    return best["hours"], best["members"]


def row_estimated_hours(row: pd.Series, revised_col: str | None = None) -> float | None:
    """
    Hours for one tracker row.

    Total Revised Estimation wins when present; otherwise Estimated Hrs.
    """
    estimated = _as_float(row.get(COL_ESTIMATED_HRS)) if COL_ESTIMATED_HRS in row.index else None
    if revised_col is None:
        revised_col = revised_estimation_column(row.index)
    if not revised_col or revised_col not in row.index:
        return estimated
    revised = _as_float(row.get(revised_col))
    if revised is not None:
        return revised
    return estimated


def sum_scrum_estimated_hours(tracker: pd.DataFrame | None) -> int | None:
    """Sum per-row estimated hours (revised when present) across the Scrum tracker."""
    if tracker is None or tracker.empty:
        return None
    revised_col = revised_estimation_column(tracker.columns)
    if COL_ESTIMATED_HRS not in tracker.columns and revised_col is None:
        return None

    total = 0.0
    saw_any = False
    for _, row in tracker.iterrows():
        hours = row_estimated_hours(row, revised_col=revised_col)
        if hours is None:
            continue
        total += hours
        saw_any = True
    if not saw_any:
        return 0
    return int(round(total))


def _sum_column(tracker: pd.DataFrame, column: str) -> float:
    if column not in tracker.columns:
        return 0.0
    total = 0.0
    for value in tracker[column]:
        hours = _as_float(value)
        if hours is not None:
            total += hours
    return total


def sum_burndown_hours(tracker: pd.DataFrame | None) -> int | None:
    """Actual Efforts (Hrs) minus Competency Gap Efforts across the tracker."""
    if tracker is None or tracker.empty:
        return None
    if COL_ACTUAL_EFFORTS not in tracker.columns and COL_COMPETENCY_GAP not in tracker.columns:
        return None
    actual = _sum_column(tracker, COL_ACTUAL_EFFORTS)
    gap = _sum_column(tracker, COL_COMPETENCY_GAP)
    return int(round(actual - gap))


def load_quarterly_planning(
    planning_book: str | Path | None = None,
    *,
    planned_pct: int = DEFAULT_PLANNED_BANDWIDTH_PCT,
    tracker: pd.DataFrame | None = None,
) -> dict[str, int] | None:
    """
    Build the three planning-bar metrics.

    ``planned_pct`` is ignored for the second bar (kept for call-site compatibility).
    """
    del planned_pct  # second bar is Scrum estimates, not % of available

    workbook_path = Path(planning_book) if planning_book else DEFAULT_PLANNING_BOOK
    if not workbook_path.exists():
        return None

    wb = load_workbook(workbook_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    available_hours, resources = _lookup_available_hours(ws)
    if available_hours is None:
        return None

    estimated_hours = sum_scrum_estimated_hours(tracker)
    if estimated_hours is None:
        estimated_hours = 0

    burndown_hours = sum_burndown_hours(tracker)
    if burndown_hours is None:
        burndown_hours = 0

    available = int(round(available_hours))
    planned_pct = int(round((estimated_hours / available) * 100)) if available else 0

    return {
        "available_hours": available,
        # Second bar value (legacy key name kept for slide/chart callers).
        "planned_hours": int(estimated_hours),
        "estimated_hours": int(estimated_hours),
        "burndown_hours": int(burndown_hours),
        "planned_pct": planned_pct,
        "resources": resources if resources is not None else 0,
    }
